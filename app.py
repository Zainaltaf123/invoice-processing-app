# app.py
import streamlit as st
import pandas as pd
import zipfile
import pdfplumber
import re
import os
import io
from openpyxl import load_workbook
from zipfile import ZipFile
import base64

st.title("Invoice Processing Tool with ZIP or Single PDF Upload")

zip_file = st.file_uploader("Upload ZIP of Invoices (Optional)", type=["zip"])
pdf_file = st.file_uploader("Or Upload a Single Invoice PDF", type=["pdf"])
paf_file = st.file_uploader("Upload PAF Excel", type=["xlsx"])
template_file = st.file_uploader("Upload Invoice Template Excel", type=["xlsx"])

temp_folder = "temp_excels"
final_folder = "final_outputs"
final_invoice_folder = "final_invoice_output"
summary_file = "Summary_Report.xlsx"

for folder in [temp_folder, final_folder, final_invoice_folder]:
    os.makedirs(folder, exist_ok=True)

line_pattern = re.compile(
    r"^\s*(\d+)\s+"            # SNo
    r"(\S+)\s+"                # Product
    r"(.+?)\s+"                # Description
    r"(\d+(?:\.?\d*)?(?:EA|PAC))\s+"  # Quantity+Unit
    r"([\d\.]+)\s+"            # Gross Price
    r"([\d\.]+)\s*$"           # Extension Cost
)

# Map full province names to their 2‑letter codes
PROVINCE_MAP = {
    'alberta': 'AB', 'british columbia': 'BC', 'manitoba': 'MB',
    'new brunswick': 'NB', 'newfoundland and labrador': 'NL', 'nova scotia': 'NS',
    'northwest territories': 'NT', 'nunavut': 'NU', 'ontario': 'ON',
    'prince edward island': 'PE', 'quebec': 'QC', 'saskatchewan': 'SK', 'yukon': 'YT'
}
CODES = list(PROVINCE_MAP.values())

def extract_shipto_province_from_text(text):
    lines = text.splitlines()
    # find “Ship To” header
    start = next((i for i,l in enumerate(lines) if "ship to" in l.lower()), None)
    if start is None:
        return None
    # find next section header
    end = next(
      (j for j in range(start+1, len(lines))
       if any(h in lines[j].lower() for h in ("customer id","invoice no","bill to"))),
      len(lines)
    )
    block_text = " ".join(lines[start+1:end]).lower()
    hits = []
    for fullname, code in PROVINCE_MAP.items():
        idx = block_text.rfind(fullname)
        if idx >= 0:
            hits.append((idx, code))
    for code in CODES:
        m = re.search(r'\b'+code.lower()+r'\b', block_text)
        if m:
            hits.append((m.start(), code))
    if not hits:
        return None
    hits.sort(key=lambda x: x[0])
    return hits[-1][1]

def extract_invoice_data(file):
    products = []
    invoice_number = pic_number = freight_charges = gst_amount = total_tax_included = order_number = ship_to_address = province = None
    skip = False

    with pdfplumber.open(file) as pdf:
        first_text = pdf.pages[0].extract_text()

        # invoice & PIC
        inv_m = re.search(r"(INV\d{6})", first_text)
        pic_m = re.search(r"(PIC\d{6})", first_text)
        invoice_number = inv_m.group(1) if inv_m else None
        pic_number = pic_m.group(1) if pic_m else None

        # order number
        for line in first_text.splitlines():
            m = re.search(r"\b(RGRHO\w+|CCAO\w+)\b", line)
            if m:
                order_number = m.group(1)
                break
        if order_number and order_number.startswith("CCAO"):
            skip = True

        # ship to address
        ship_to_lines = re.findall(r"Ship To\s*\n(.*?)\n(.*?)\n", first_text, re.DOTALL)
        ship_to_address = "\n".join(ship_to_lines[0]) if ship_to_lines else ""
        # province via helper
        province = extract_shipto_province_from_text(first_text)

        # line items
        for page in pdf.pages:
            lines = page.extract_text().split("\n")
            try:
                header_idx = next(i for i,line in enumerate(lines)
                                  if line.strip().startswith("SNo.") and "Extension Cost" in line)
            except StopIteration:
                continue
            for line in lines[header_idx+1:]:
                if line.strip().startswith("Page"):
                    break
                m = line_pattern.match(line)
                if m:
                    qty_full = m.group(4)
                    qty_num = re.match(r"([\d\.]+)", qty_full).group(1)
                    qty_unit = re.search(r"(EA|PAC)", qty_full).group(1)
                    products.append({
                        "SNo": int(m.group(1)),
                        "Product": m.group(2),
                        "Description": m.group(3).strip(),
                        "Quantity": float(qty_num),
                        "Unit": qty_unit,
                        "Gross Price": float(m.group(5)),
                        "Extension Cost": float(m.group(6))
                    })

        # footer values
        last_text = pdf.pages[-1].extract_text()
        f_m = re.search(r"Freight Charges\s+([\d,]+\.\d{2})", last_text)
        g_m = re.search(r"GST/HST Amount\s+([\d,]+\.\d{2})", last_text)
        t_m = re.search(r"TOTAL TAX INCLUDED\s+([\d,]+\.\d{2})", last_text, re.IGNORECASE)
        freight_charges = float(f_m.group(1).replace(",","")) if f_m else 0
        gst_amount = float(g_m.group(1).replace(",","")) if g_m else 0
        total_tax_included = float(t_m.group(1).replace(",","")) if t_m else 0

    df = pd.DataFrame(products)
    return (df, invoice_number, pic_number, freight_charges,
            gst_amount, total_tax_included, order_number,
            ship_to_address, province, skip)

if (zip_file or pdf_file) and paf_file and template_file:
    if st.button("Process Invoices"):

        # cleanup
        for folder in [temp_folder, final_folder, final_invoice_folder]:
            for f in os.listdir(folder):
                os.remove(os.path.join(folder, f))

        # save uploads
        with open("uploaded_paf.xlsx","wb") as f: f.write(paf_file.getbuffer())
        with open("uploaded_template.xlsx","wb") as f: f.write(template_file.getbuffer())

        paf_df = pd.read_excel("uploaded_paf.xlsx")
        paf_df.columns = paf_df.columns.str.strip()
        paf_df = paf_df.drop_duplicates(subset=["Valiant/RGR SKU"])

        # collect pdf sources
        pdf_sources = []
        if zip_file:
            with open("uploaded_invoices.zip","wb") as f: f.write(zip_file.getbuffer())
            with zipfile.ZipFile("uploaded_invoices.zip",'r') as zip_ref:
                pdf_sources = [(f, zip_ref.open(f)) for f in zip_ref.namelist() if f.lower().endswith('.pdf')]
        else:
            pdf_sources = [(pdf_file.name, pdf_file)]

        total_files = len(pdf_sources)
        summary_list = []
        missing_products_list = []
        progress_bar = st.progress(0)
        status_text = st.empty()

        for i,(filename,pdf_stream) in enumerate(pdf_sources, start=1):
            status_text.text(f"Processing {filename} ({i}/{total_files})...")
            (invoice_df, invoice_number, pic_number, freight,
             gst, total_tax, order_number,
             ship_to, province, skip) = extract_invoice_data(pdf_stream)
            base_name = os.path.splitext(os.path.basename(filename))[0]

            # save raw
            invoice_df.to_excel(os.path.join(temp_folder,f"{base_name}_raw_invoice.xlsx"),index=False)

            if skip:
                summary_list.append({
                    "Invoice File": base_name,
                    "Invoice Number": invoice_number,
                    "PIC Number": pic_number,
                    "Order Number": order_number,
                    "Ship To": ship_to,
                    "Province": province,
                    "Products in Invoice": 0,
                    "Products Matched to PAF": 0,
                    "Total Tax Included (Original)": total_tax,
                    "Calculated Total Payable": 0
                })
                continue

            merged_df = pd.merge(invoice_df, paf_df, left_on="Product", right_on="Valiant/RGR SKU", how="left")
            merged_df["Units Per Case"] = pd.to_numeric(merged_df["Units Per Case"],errors="coerce")
            merged_df["Quantity"] = pd.to_numeric(merged_df["Quantity"],errors="coerce")
            merged_df["Total Quantity"] = merged_df["Quantity"] * merged_df["Units Per Case"]
            merged_df["Unit Cost Price"] = merged_df["Gross Price"] / merged_df["Units Per Case"]

            final_df = merged_df[["GlobalTill SKU","Total Quantity","Unit Cost Price"]].copy()
            final_df.columns = ["SKU","Total Quantity","Unit Cost Price"]

            wb = load_workbook("uploaded_template.xlsx")
            ws = wb.active

            # header values
            ws["B6"].value = freight
            ws["B7"].value = "rgr canada"
            ws["B8"].value = f"{invoice_number}/{pic_number}"
            ws["B9"].value = 0

            # GST placement based on province
            if province == "ON":
                ws["B10"].value = 0
                ws["B11"].value = gst
            else:
                ws["B10"].value = gst
                ws["B11"].value = 0

            # line items
            start_row = 14
            for idx,prd in final_df.iterrows():
                ws[f"A{start_row+idx}"].value = prd["SKU"]
                ws[f"B{start_row+idx}"].value = prd["Total Quantity"]
                ws[f"C{start_row+idx}"].value = prd["Unit Cost Price"]

            wb.save(os.path.join(final_invoice_folder,f"{base_name}_final_invoice.xlsx"))

            # summary calculations
            product_count_invoice = len(invoice_df)
            product_count_processed = final_df["SKU"].notna().sum()
            calculated_total = (merged_df["Total Quantity"]*merged_df["Unit Cost Price"]).sum() + gst + freight

            summary_list.append({
                "Invoice File": base_name,
                "Invoice Number": invoice_number,
                "PIC Number": pic_number,
                "Order Number": order_number,
                "Ship To": ship_to,
                "Province": province,
                "Products in Invoice": product_count_invoice,
                "Products Matched to PAF": product_count_processed,
                "Total Tax Included (Original)": total_tax,
                "Calculated Total Payable": round(calculated_total,2)
            })

            # missing products
            for _,row in merged_df[merged_df["GlobalTill SKU"].isna()].iterrows():
                missing_products_list.append({
                    "Invoice File": base_name,
                    "Product": row["Product"],
                    "Description": row["Description"],
                    "Quantity": row["Quantity"],
                    "Gross Price": row["Gross Price"]
                })

            progress_bar.progress(i/total_files)

        # write summary Excel
        summary_df = pd.DataFrame(summary_list)
        missing_products_df = pd.DataFrame(missing_products_list)
        paf_data_df = paf_df

        with pd.ExcelWriter(summary_file, engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name="Summary Report", index=False)
            missing_products_df.to_excel(writer, sheet_name="Missing Products", index=False)
            paf_data_df.to_excel(writer, sheet_name="PAF Data", index=False)

        st.success("✅ Processing Complete!")

        # download links
        with open(summary_file,"rb") as f:
            b64=base64.b64encode(f.read()).decode()
            st.markdown(f'<a href="data:application/octet-stream;base64,{b64}" download="invoice_summary.xlsx">📥 Download Invoice Summary</a>', unsafe_allow_html=True)

        st.write("### Download Final Invoices")
        for fname in os.listdir(final_invoice_folder):
            if fname.endswith(".xlsx"):
                with open(os.path.join(final_invoice_folder,fname),"rb") as f:
                    st.download_button(f"Download {fname}", f, fname)

        # zip all final invoices
        buf=io.BytesIO()
        with ZipFile(buf,"w") as z:
            for fname in os.listdir(final_invoice_folder):
                if fname.endswith(".xlsx"):
                    z.write(os.path.join(final_invoice_folder,fname), fname)
        buf.seek(0)
        st.download_button("Download All Final Invoices as ZIP", buf, "final_invoices.zip")
